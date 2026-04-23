"""
Exporta inventario AWS a Excel.

Genera un workbook con una hoja resumen y una hoja por servicio
consolidando la informacion cacheada por cuenta y region.
"""

from datetime import datetime
import logging

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E2E", end_color="1F4E2E", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

SERVICE_EXPORT_CONFIG = [
    {
        "cache_key": "ec2",
        "sheet_name": "EC2",
        "summary_label": "EC2",
        "global_service": False,
        "preferred_columns": ["cuenta", "region", "nombre", "id", "tipo", "estado", "vpc", "ip_privada", "ip_publica"],
    },
    {
        "cache_key": "rds",
        "sheet_name": "RDS",
        "summary_label": "RDS",
        "global_service": False,
        "preferred_columns": ["cuenta", "region", "nombre", "id", "motor", "version", "estado", "tipo", "almacenamiento_gb"],
    },
    {
        "cache_key": "vpc",
        "sheet_name": "VPC",
        "summary_label": "VPC",
        "global_service": False,
        "preferred_columns": ["cuenta", "region", "nombre", "id", "cidr", "estado", "subnets"],
    },
    {
        "cache_key": "vpc_outbound_ips",
        "sheet_name": "NAT_IPs",
        "summary_label": "NAT/IPs",
        "global_service": False,
        "preferred_columns": ["cuenta", "region", "type", "name", "resource_id", "public_ip", "private_ip", "vpc_id", "subnet_id", "state"],
    },
    {
        "cache_key": "s3",
        "sheet_name": "S3",
        "summary_label": "S3",
        "global_service": True,
        "preferred_columns": ["cuenta", "nombre", "region", "creacion"],
    },
    {
        "cache_key": "iam_users",
        "sheet_name": "IAM_Users",
        "summary_label": "IAM Users",
        "global_service": True,
        "preferred_columns": ["cuenta", "username", "arn", "mfa_enabled", "access_keys", "creacion"],
    },
    {
        "cache_key": "lambda",
        "sheet_name": "Lambda",
        "summary_label": "Lambda",
        "global_service": False,
        "preferred_columns": [
            "cuenta", "region", "nombre", "handler", "runtime", "memoria_mb", "timeout_s",
            "estado", "vpc", "subnets", "security_groups", "execution_role_name",
            "access_actions", "access_resources"
        ],
    },
    {
        "cache_key": "api_gateway",
        "sheet_name": "API_Gateway",
        "summary_label": "API Gateway",
        "global_service": False,
        "preferred_columns": [
            "cuenta", "region", "nombre", "tipo", "estado", "rutas",
            "integraciones_lambda", "lambdas"
        ],
    },
    {
        "cache_key": "api_gateway_routes",
        "sheet_name": "API_Lambda_Map",
        "summary_label": "API Gateway Routes",
        "global_service": False,
        "preferred_columns": [
            "cuenta", "region", "api_nombre", "api_id", "api_tipo", "route_key",
            "metodo_http", "ruta", "integration_type", "lambda_function", "lambda_handler",
            "lambda_runtime", "lambda_execution_role", "lambda_vpc", "lambda_subnets",
            "lambda_security_groups", "lambda_access_actions", "lambda_access_resources"
        ],
    },
    {
        "cache_key": "cloudformation",
        "sheet_name": "CloudFormation",
        "summary_label": "CloudFormation",
        "global_service": False,
        "preferred_columns": ["cuenta", "region", "nombre", "estado"],
    },
    {
        "cache_key": "ssm",
        "sheet_name": "SSM",
        "summary_label": "SSM",
        "global_service": False,
        "preferred_columns": ["cuenta", "region", "nombre", "tipo", "tier", "version", "data_type"],
    },
    {
        "cache_key": "kms",
        "sheet_name": "KMS",
        "summary_label": "KMS",
        "global_service": False,
        "preferred_columns": ["cuenta", "region", "alias", "key_id", "arn", "estado", "manager", "origen"],
    },
    {
        "cache_key": "dynamodb",
        "sheet_name": "DynamoDB",
        "summary_label": "DynamoDB",
        "global_service": False,
        "preferred_columns": ["cuenta", "region", "nombre", "estado", "billing_mode", "lectura", "escritura"],
    },
    {
        "cache_key": "sqs",
        "sheet_name": "SQS",
        "summary_label": "SQS",
        "global_service": False,
        "preferred_columns": ["cuenta", "region", "nombre", "fifo", "kms_key_id"],
    },
]


def _remove_timezones(df):
    """Convierte columnas datetime timezone-aware a valores compatibles con Excel."""
    if df.empty:
        return df

    clean_df = df.copy()
    for column in clean_df.columns:
        try:
            if pd.api.types.is_datetime64_any_dtype(clean_df[column]):
                if hasattr(clean_df[column].dtype, "tz") and clean_df[column].dtype.tz is not None:
                    clean_df[column] = clean_df[column].dt.tz_convert("UTC").dt.tz_localize(None)
                elif clean_df[column].dt.tz is not None:
                    clean_df[column] = clean_df[column].dt.tz_localize(None)
        except Exception as exc:
            logger.warning("Advertencia limpiando %s: %s", column, exc)
            try:
                clean_df[column] = clean_df[column].astype(str)
            except Exception:
                continue

    return clean_df


def _apply_styles(ws, df):
    """Aplica estilos base y autoajuste de columnas."""
    for col_num, colname in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = colname
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col_num, colname in enumerate(df.columns, 1):
        max_length = max(
            df.iloc[:, col_num - 1].astype(str).map(len).max() if not df.empty else 0,
            len(str(colname)),
        )
        ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 50)


def _get_account_regions(perfiles, account):
    """Obtiene las regiones configuradas para una cuenta."""
    return perfiles.get(account, {}).get("regiones", []) or ["us-east-1"]


def _get_global_region(perfiles, account):
    """Obtiene la region base para servicios globales."""
    account_config = perfiles.get(account, {})
    return account_config.get("region") or _get_account_regions(perfiles, account)[0]


def _reorder_dataframe_columns(df, preferred_columns):
    """Reordena columnas dejando primero las mas utiles."""
    existing_cols = [column for column in preferred_columns if column in df.columns]
    return df[existing_cols + [column for column in df.columns if column not in existing_cols]]


def _aggregate_service(cache_manager, accounts, perfiles, service_config):
    """Consolida un servicio para una o multiples cuentas."""
    dfs = []

    for account in accounts:
        regions = [_get_global_region(perfiles, account)] if service_config["global_service"] else _get_account_regions(perfiles, account)

        for region in regions:
            try:
                data, _, exists = cache_manager.get(account, region, service_config["cache_key"])
            except Exception:
                continue

            if not (exists and isinstance(data, pd.DataFrame) and not data.empty):
                continue

            service_df = data.copy()
            if "cuenta" not in service_df.columns:
                service_df["cuenta"] = account
            if "region" not in service_df.columns:
                service_df["region"] = region
            dfs.append(service_df)

    if not dfs:
        return pd.DataFrame()

    merged_df = pd.concat(dfs, ignore_index=True)
    return _reorder_dataframe_columns(merged_df, service_config["preferred_columns"])


def export_to_excel(cache_manager, accounts, perfiles, output_path="inventario_aws.xlsx"):
    """
    Exporta el inventario cacheado a un archivo Excel.

    Args:
        cache_manager: Instancia con metodo get(account, region, service_key)
        accounts: Lista de cuentas a incluir
        perfiles: Configuracion de perfiles/regiones
        output_path: Ruta de salida del workbook

    Returns:
        Ruta generada o None si ocurre un error
    """
    try:
        logger.info("Generando Excel: %s", output_path)

        export_data = {}
        summary_rows = []

        for service_config in SERVICE_EXPORT_CONFIG:
            service_df = _remove_timezones(_aggregate_service(cache_manager, accounts, perfiles, service_config))
            export_data[service_config["cache_key"]] = service_df
            summary_rows.append(
                {
                    "Servicio": service_config["summary_label"],
                    "Cantidad": len(service_df),
                }
            )

        wb = Workbook()
        wb.remove(wb.active)

        ws_summary = wb.create_sheet("Resumen", 0)
        ws_summary["A1"] = "INVENTARIO AWS"
        ws_summary["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws_summary["A1"].fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        ws_summary.merge_cells("A1:C1")

        ws_summary["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary["A2"].font = Font(italic=True, size=10)

        ws_summary["A4"] = "Cuenta(s)"
        ws_summary["B4"] = ", ".join(accounts)
        ws_summary["A5"] = "Regiones"
        ws_summary["B5"] = "Todas las configuradas para cada cuenta"

        summary_df = pd.DataFrame(summary_rows)
        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 7):
            for c_idx, value in enumerate(row, 1):
                ws_summary.cell(row=r_idx, column=c_idx, value=value)

        for header_cell in ("A4", "A5"):
            ws_summary[header_cell].font = Font(bold=True)

        header_row = 7
        for col_num in range(1, len(summary_df.columns) + 1):
            cell = ws_summary.cell(row=header_row, column=col_num)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

        for row_num in range(header_row + 1, header_row + 1 + len(summary_df)):
            for col_num in range(1, len(summary_df.columns) + 1):
                cell = ws_summary.cell(row=row_num, column=col_num)
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="left", vertical="top")

        total_row = header_row + len(summary_df) + 1
        ws_summary[f"A{total_row}"] = "TOTAL"
        ws_summary[f"B{total_row}"] = int(summary_df["Cantidad"].sum())
        ws_summary[f"A{total_row}"].font = Font(bold=True)
        ws_summary[f"B{total_row}"].font = Font(bold=True)

        for column in ("A", "B", "C"):
            ws_summary.column_dimensions[column].width = 28

        sheet_index = 1
        for service_config in SERVICE_EXPORT_CONFIG:
            service_df = export_data[service_config["cache_key"]]
            if service_df.empty:
                continue

            ws_service = wb.create_sheet(service_config["sheet_name"], sheet_index)
            for r_idx, row in enumerate(dataframe_to_rows(service_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_service.cell(row=r_idx, column=c_idx, value=value)

            _apply_styles(ws_service, service_df)
            logger.info("Pestana %s: %s filas", service_config["sheet_name"], len(service_df))
            sheet_index += 1

        wb.save(output_path)
        logger.info("Excel guardado: %s", output_path)
        return output_path
    except Exception as exc:
        logger.error("Error generando Excel: %s", exc)
        return None
